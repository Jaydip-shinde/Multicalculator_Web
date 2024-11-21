from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = "form_data.xlsx"

# Ensure the Excel file exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Form Responses"
    ws.append(["Name", "Message Type", "Message"])  # Add headers
    wb.save(EXCEL_FILE)

@app.route('/submit', methods=['POST'])
def submit_form():
    data = request.form
    name = data.get('name')
    message_type = data.get('type')  # contact, feedback, queries, suggestions
    message = data.get('message')

    if not all([name, message_type, message]):
        return jsonify({"status": "error", "message": "All fields are required"}), 400

    # Save data to Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, message_type, message])
    wb.save(EXCEL_FILE)

    return jsonify({"status": "success", "message": "Form data saved successfully!"})

if __name__ == '__main__':
    app.run(debug=True)

def run_script():
    # Place your script logic here
    try:
        # For example, log a message or perform a task
        print("The script is running automatically!")
        return {"status": "success", "message": "Script executed successfully"}
    except Exception as e:
        return {"status": "error", "message": str(e)}   
